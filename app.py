import os
import time
import re
import json
import random
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import zipfile

# Load environment variables
load_dotenv(override=True)

app = Flask(__name__, static_folder='public', static_url_path='')
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_PDF_EXTENSIONS = {'pdf'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB for PDFs
MAX_EXCEL_SIZE = 25 * 1024 * 1024  # 25MB for Excel

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Create uploads directory if it doesn't exist
Path(UPLOAD_FOLDER).mkdir(exist_ok=True)

# In-memory storage
document_analysis = {}  # {fileId: {filename, content, tables, uploadTime}}
file_paths = {}  # {fileId: filePath}
excel_templates = {}  # {excelId: {filename, path, uploadTime}}
filled_workbooks = {}  # {filledId: {filename, path, createdTime}}
fill_jobs = {}  # {jobId: {status, current, total, startTime, error, result}}


# Utility functions
def get_env(name):
    """Get environment variable and trim whitespace."""
    return (os.getenv(name) or '').strip()


def is_placeholder(value):
    """Check if value is a placeholder."""
    if not value:
        return True
    return bool(re.match(r'^your_', value, re.I)) or bool(re.search(r'_here$', value, re.I))


def require_env(name):
    """Require environment variable to be set and not a placeholder."""
    value = get_env(name)
    if is_placeholder(value):
        raise ValueError(f'Missing or placeholder {name}. Update your .env file.')
    return value


def require_any_env(names, canonical_name_for_error=None):
    """Require at least one environment variable from a list."""
    for name in names:
        value = get_env(name)
        if not is_placeholder(value):
            return value
    canonical = canonical_name_for_error or names[0] if names else 'ENV_VAR'
    raise ValueError(f'Missing or placeholder {canonical}. Update your .env file.')


def normalize_base_url(raw_url, name_for_errors):
    """Normalize base URL to ensure it ends with /."""
    try:
        if not raw_url.startswith(('http://', 'https://')):
            raise ValueError('Unsupported protocol')
        if not raw_url.endswith('/'):
            raw_url = raw_url + '/'
        return raw_url
    except Exception as e:
        raise ValueError(
            f'Invalid {name_for_errors}. Expected a full URL like https://<your-resource>/'
        )


def create_id(prefix='id'):
    """Create a unique ID."""
    return f"{prefix}-{int(time.time() * 1000)}-{os.urandom(4).hex()}"


def sanitize_filename(name):
    """Sanitize filename for safe storage."""
    return re.sub(r'[^a-zA-Z0-9._-]', '_', str(name or 'file'))


def is_zip_file_header(file_path):
    """Check if file has ZIP header (for .xlsx validation)."""
    try:
        with open(file_path, 'rb') as f:
            header = f.read(4)
            if len(header) < 4:
                return False
            # ZIP files start with: 50 4B 03 04
            return header[0] == 0x50 and header[1] == 0x4B
    except:
        return False


def allowed_file(filename, extensions):
    """Check if filename has allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions


def rebuild_file_paths_mapping():
    """Rebuild file paths mapping from existing files."""
    try:
        upload_dir = Path(UPLOAD_FOLDER)
        if upload_dir.exists():
            for file_path in upload_dir.iterdir():
                if file_path.is_file():
                    file_paths[file_path.name] = str(file_path)
            print(f'Rebuilt file paths mapping for {len(file_paths)} files')
    except Exception as error:
        print(f'Error rebuilding file paths mapping: {error}')


# Azure Content Understanding integration
def analyze_pdf(file_path, original_filename, retry_count=0):
    """Analyze PDF using Azure Document Intelligence."""
    max_retries = 3
    base_delay = 1  # 1 second base delay
    
    try:
        endpoint_base = normalize_base_url(
            require_env('AZURE_CONTENT_UNDERSTANDING_ENDPOINT'),
            'AZURE_CONTENT_UNDERSTANDING_ENDPOINT'
        )
        subscription_key = require_env('AZURE_CONTENT_UNDERSTANDING_KEY')
        
        analyze_url = urljoin(
            endpoint_base,
            'formrecognizer/documentModels/prebuilt-layout:analyze?api-version=2023-07-31'
        )
        
        # Start document analysis
        with open(file_path, 'rb') as f:
            files = {'file': (original_filename, f, 'application/pdf')}
            headers = {'Ocp-Apim-Subscription-Key': subscription_key}
            
            analyze_response = requests.post(
                analyze_url,
                headers=headers,
                files=files,
                timeout=90
            )
            analyze_response.raise_for_status()
        
        # Get operation location
        operation_location = analyze_response.headers.get('operation-location')
        if not operation_location:
            raise Exception('No operation location received from Content Understanding API')
        
        # Poll for results
        attempts = 0
        max_attempts = 30  # 30 attempts * 2 seconds = 1 minute timeout
        
        while attempts < max_attempts:
            time.sleep(2)  # Wait 2 seconds
            result_response = requests.get(
                operation_location,
                headers={'Ocp-Apim-Subscription-Key': subscription_key}
            )
            result_response.raise_for_status()
            result_data = result_response.json()
            
            if result_data.get('status') == 'succeeded':
                # Extract text and tables
                analyze_result = result_data.get('analyzeResult', {})
                pages = analyze_result.get('pages', [])
                tables = analyze_result.get('tables', [])
                extracted_text = ''
                
                # Extract text from pages
                for page in pages:
                    if page.get('lines'):
                        for line in page['lines']:
                            extracted_text += line.get('content', '') + '\n'
                
                # Extract tables
                extracted_tables = []
                if tables:
                    extracted_text += '\n\n=== TABLES ===\n\n'
                    
                    for table_index, table in enumerate(tables):
                        row_count = table.get('rowCount', 0)
                        column_count = table.get('columnCount', 0)
                        extracted_text += f"\n--- Table {table_index + 1} ({row_count} rows x {column_count} columns) ---\n"
                        
                        # Create 2D array for table
                        table_data = [['' for _ in range(column_count)] for _ in range(row_count)]
                        
                        # Fill in cell values
                        for cell in table.get('cells', []):
                            row_index = cell.get('rowIndex', 0)
                            col_index = cell.get('columnIndex', 0)
                            table_data[row_index][col_index] = cell.get('content', '')
                        
                        # Store structured table data
                        extracted_tables.append({
                            'tableIndex': table_index + 1,
                            'rowCount': row_count,
                            'columnCount': column_count,
                            'data': table_data
                        })
                        
                        # Format as text table
                        for row in table_data:
                            extracted_text += ' | '.join(row) + '\n'
                        extracted_text += '\n'
                
                return {'text': extracted_text.strip(), 'tables': extracted_tables}
            
            elif result_data.get('status') != 'running':
                raise Exception(f"Document analysis failed with status: {result_data.get('status')}")
            
            attempts += 1
        
        raise Exception('Document analysis timed out')
        
    except Exception as error:
        print(f"Content Understanding API Error (attempt {retry_count + 1}): {error}")
        
        # Configuration/URL issues
        if 'Missing or placeholder AZURE_CONTENT_UNDERSTANDING_' in str(error):
            raise
        if 'Invalid AZURE_CONTENT_UNDERSTANDING_ENDPOINT' in str(error):
            raise
        
        # Handle rate limiting with exponential backoff
        if hasattr(error, 'response') and error.response and error.response.status_code == 429:
            if retry_count < max_retries:
                delay = base_delay * (2 ** retry_count) + (random.random() * 1)
                print(f'Rate limited. Retrying in {round(delay)}ms... (attempt {retry_count + 1}/{max_retries})')
                time.sleep(delay)
                return analyze_pdf(file_path, original_filename, retry_count + 1)
            raise Exception(f'Rate limit exceeded. Please wait a few minutes before uploading more files. (Tried {retry_count + 1} times)')
        
        # User-friendly error messages
        if hasattr(error, 'response') and error.response:
            status_code = error.response.status_code
            if status_code == 401:
                raise Exception('Authentication failed - please check your API key')
            elif status_code == 413:
                raise Exception('File too large for Content Understanding service')
            elif status_code == 400:
                raise Exception('Invalid PDF file format or corrupted file')
        
        raise Exception(f'Document analysis failed: {error}')


# Azure OpenAI integration
def truncate_section_text(text, max_chars=16000):
    """Truncate text to max_chars, keeping head and tail."""
    safe = str(text or '')
    if len(safe) <= max_chars:
        return safe
    head = safe[:max_chars // 2]
    tail = safe[-(max_chars // 2):]
    return f"{head}\n\n[...TRUNCATED...]\n\n{tail}"


def extract_json_object(text):
    """Extract JSON object from text."""
    raw = str(text or '').strip()
    if not raw:
        return None
    
    try:
        return json.loads(raw)
    except:
        pass
    
    # Try to find JSON object in text
    start = raw.find('{')
    end = raw.rfind('}')
    if start >= 0 and end > start:
        candidate = raw[start:end + 1]
        try:
            return json.loads(candidate)
        except:
            return None
    
    return None


def normalize_gpt_answer_payload(payload, marker):
    """Normalize GPT answer payload."""
    if not payload or not isinstance(payload, dict):
        return None
    
    answer = payload.get('answer', '').strip() if isinstance(payload.get('answer'), str) else None
    confidence_raw = payload.get('confidence')
    
    try:
        confidence = float(confidence_raw) if confidence_raw is not None else float('nan')
    except:
        confidence = float('nan')
    
    evidence = payload.get('evidence_snippet', '').strip() if isinstance(payload.get('evidence_snippet'), str) else ''
    marker_out = payload.get('marker', '').strip() if isinstance(payload.get('marker'), str) else marker
    
    if not answer or not (0 <= confidence <= 1 or confidence != confidence):  # Check for NaN
        return None
    
    bounded_confidence = max(0, min(1, confidence))
    
    return {
        'answer': answer[:250] if len(answer) > 250 else answer,
        'confidence': bounded_confidence,
        'evidence_snippet': evidence,
        'marker': marker_out or marker
    }


def generate_answer_with_azure_openai(marker, question, section_text):
    """Generate answer using Azure OpenAI."""
    endpoint_base = normalize_base_url(
        require_any_env(
            ['AZURE_OPENAI_ENDPOINT', 'REACT_APP_AZURE_OPENAI_CHAT_ENDPOINT', 'REACT_APP_AZURE_OPENAI_ENDPOINT'],
            'AZURE_OPENAI_ENDPOINT'
        ),
        'AZURE_OPENAI_ENDPOINT'
    )
    api_key = require_any_env(
        ['AZURE_OPENAI_API_KEY', 'AZURE_OPENAI_KEY', 'REACT_APP_AZURE_OPENAI_CHAT_KEY', 'REACT_APP_AZURE_OPENAI_KEY'],
        'AZURE_OPENAI_API_KEY'
    )
    deployment = require_any_env(
        ['AZURE_OPENAI_DEPLOYMENT', 'REACT_APP_AZURE_OPENAI_DEPLOYMENT', 'REACT_APP_AZURE_OPENAI_DEPLOYMENT_NAME'],
        'AZURE_OPENAI_DEPLOYMENT'
    )
    api_version = get_env('AZURE_OPENAI_API_VERSION') or '2024-02-15-preview'
    
    url = urljoin(
        endpoint_base,
        f"openai/deployments/{deployment}/chat/completions?api-version={api_version}"
    )
    
    safe_marker = str(marker or '').strip()
    safe_question = str(question or '').strip()
    safe_section_text = truncate_section_text(section_text or '')
    
    system_content = (
        'You are a helpful assistant. Answer questions based on the following document content:\n\n'
        f'{safe_section_text}\n\n'
        'All questions are in column A, all answers should be placed in Column B in the cell to the right. '
        'If the answer cannot be found respond "N/A". \n\n'
        'Return ONLY a single valid JSON object with EXACT keys: answer (string <= 250 chars), '
        'confidence (number 0-1), evidence_snippet (string), marker (string). '
        'Do not include any extra keys or text.'
    )
    
    user_content = (
        f'marker: {safe_marker}\n'
        f'question: {safe_question}\n'
        f'section_text:\n{safe_section_text}'
    )
    
    max_attempts = 3
    for attempt in range(1, max_attempts + 1):
        try:
            response = requests.post(
                url,
                headers={
                    'api-key': api_key,
                    'Content-Type': 'application/json'
                },
                json={
                    'messages': [
                        {'role': 'system', 'content': system_content},
                        {'role': 'user', 'content': user_content}
                    ],
                    'max_tokens': 800,
                    'temperature': 0.7,
                    'response_format': {'type': 'json_object'}
                },
                timeout=90
            )
            
            if response.status_code == 200:
                content = response.json().get('choices', [{}])[0].get('message', {}).get('content', '')
                parsed = extract_json_object(content)
                normalized = normalize_gpt_answer_payload(parsed, safe_marker)
                if normalized:
                    return normalized
                
                if attempt == max_attempts:
                    return {
                        'answer': 'NOT FOUND',
                        'confidence': 0,
                        'evidence_snippet': '',
                        'marker': safe_marker
                    }
            
            # If response_format isn't supported, retry without it
            if response.status_code == 400 and 'response_format' in response.text and attempt < max_attempts:
                try:
                    response = requests.post(
                        url,
                        headers={
                            'api-key': api_key,
                            'Content-Type': 'application/json'
                        },
                        json={
                            'messages': [
                                {'role': 'system', 'content': system_content},
                                {'role': 'user', 'content': user_content}
                            ],
                            'max_tokens': 800,
                            'temperature': 0.7
                        },
                        timeout=90
                    )
                    
                    if response.status_code == 200:
                        content = response.json().get('choices', [{}])[0].get('message', {}).get('content', '')
                        parsed = extract_json_object(content)
                        normalized = normalize_gpt_answer_payload(parsed, safe_marker)
                        if normalized:
                            return normalized
                except:
                    pass
            
            if attempt == max_attempts:
                error_msg = response.json().get('error', {}).get('message', response.text)
                raise Exception(f'Azure OpenAI answer generation failed: {error_msg}')
                
        except Exception as error:
            if attempt == max_attempts:
                raise Exception(f'Azure OpenAI answer generation failed: {error}')
    
    return {
        'answer': 'NOT FOUND',
        'confidence': 0,
        'evidence_snippet': '',
        'marker': str(marker or '').strip()
    }


# Flask Routes
@app.route('/')
def index():
    """Serve the main HTML page."""
    return send_from_directory('public', 'index.html')


@app.route('/config-status')
def config_status():
    """Debug helper: shows whether required env vars are present."""
    try:
        def explain(value):
            trimmed = str(value or '').strip()
            return {
                'set': bool(trimmed),
                'placeholder': is_placeholder(trimmed),
                'trimmedLength': len(trimmed),
                'startsWithYour_': bool(re.match(r'^your_', trimmed, re.I)),
                'endsWith_Here': bool(re.search(r'_here$', trimmed, re.I))
            }
        
        def check(name):
            return explain(get_env(name))
        
        status = {
            'AZURE_CONTENT_UNDERSTANDING_ENDPOINT': check('AZURE_CONTENT_UNDERSTANDING_ENDPOINT'),
            'AZURE_CONTENT_UNDERSTANDING_KEY': check('AZURE_CONTENT_UNDERSTANDING_KEY'),
            'AZURE_OPENAI_ENDPOINT': check('AZURE_OPENAI_ENDPOINT'),
            'AZURE_OPENAI_API_KEY': check('AZURE_OPENAI_API_KEY'),
            'AZURE_OPENAI_DEPLOYMENT': check('AZURE_OPENAI_DEPLOYMENT'),
            'AZURE_OPENAI_API_VERSION': {
                'set': bool(get_env('AZURE_OPENAI_API_VERSION')),
                'placeholder': False,
                'length': len(get_env('AZURE_OPENAI_API_VERSION'))
            }
        }
        
        ok = (
            status['AZURE_CONTENT_UNDERSTANDING_ENDPOINT']['set'] and 
            not status['AZURE_CONTENT_UNDERSTANDING_ENDPOINT']['placeholder'] and
            status['AZURE_CONTENT_UNDERSTANDING_KEY']['set'] and 
            not status['AZURE_CONTENT_UNDERSTANDING_KEY']['placeholder']
        )
        
        return jsonify({
            'ok': ok,
            'dotenv': {
                'path': str(Path('.env').absolute()),
                'exists': Path('.env').exists()
            },
            'status': status
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/pdf/<file_id>')
def serve_pdf(file_id):
    """Serve PDF files."""
    try:
        # Try to serve from local file system
        file_path = file_paths.get(file_id)
        
        # If not in mapping, try to construct path directly
        if not file_path:
            potential_path = Path(UPLOAD_FOLDER) / file_id
            if potential_path.exists():
                file_path = str(potential_path)
                file_paths[file_id] = file_path
        
        # If found, serve it
        if file_path and Path(file_path).exists():
            print(f'Serving PDF from local file: {file_path}')
            return send_file(file_path, mimetype='application/pdf', as_attachment=False)
        
        # File not found
        print(f'PDF file not found: {file_id}')
        return jsonify({'error': 'PDF file not found'}), 404
        
    except Exception as error:
        print(f'Error serving PDF: {error}')
        return jsonify({'error': 'Error retrieving PDF file'}), 500


@app.route('/upload', methods=['POST'])
def upload_file():
    """Upload and analyze PDF file."""
    try:
        if 'pdfFile' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['pdfFile']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename, ALLOWED_PDF_EXTENSIONS):
            return jsonify({'error': 'Only .pdf files are allowed'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        file_id = f"{int(time.time() * 1000)}-{filename}"
        file_path = Path(UPLOAD_FOLDER) / file_id
        file.save(file_path)
        
        # Analyze PDF
        result = analyze_pdf(str(file_path), filename)
        
        # Store analysis result
        document_analysis[file_id] = {
            'filename': filename,
            'content': result['text'],
            'tables': result['tables'],
            'uploadTime': datetime.now().isoformat()
        }
        
        # Store file path
        file_paths[file_id] = str(file_path)
        
        return jsonify({
            'success': True,
            'fileId': file_id,
            'filename': filename,
            'content': result['text']
        })
        
    except Exception as error:
        print(f'Upload error: {error}')
        return jsonify({'error': f'Upload failed: {error}'}), 500


@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    """Upload Excel workbook template."""
    try:
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No Excel file uploaded'}), 400
        
        file = request.files['excelFile']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename, ALLOWED_EXCEL_EXTENSIONS):
            return jsonify({'error': 'Only .xlsx files are allowed'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        file_id = f"{int(time.time() * 1000)}-{filename}"
        file_path = Path(UPLOAD_FOLDER) / file_id
        file.save(file_path)
        
        # Validate it's a real .xlsx file
        if not is_zip_file_header(str(file_path)):
            file_path.unlink(missing_ok=True)
            return jsonify({
                'error': (
                    'Invalid workbook file. This does not appear to be a real .xlsx (ZIP-based) file. '
                    'It may be an older .xls workbook renamed to .xlsx. '
                    'Please open it in Excel and "Save As" .xlsx, then upload again.'
                )
            }), 400
        
        excel_id = create_id('excel')
        excel_templates[excel_id] = {
            'filename': filename,
            'path': str(file_path),
            'uploadTime': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'excelId': excel_id,
            'filename': filename
        })
        
    except Exception as error:
        print(f'Excel upload error: {error}')
        return jsonify({'error': f'Excel upload failed: {error}'}), 500


@app.route('/fill-excel', methods=['POST'])
def fill_excel():
    """Start filling uploaded workbook with extracted content (background job)."""
    try:
        data = request.get_json()
        excel_id = data.get('excelId')
        file_id = data.get('fileId')
        
        if not excel_id or excel_id not in excel_templates:
            return jsonify({'error': 'Invalid or missing excelId'}), 400
        
        if not file_id or file_id not in document_analysis:
            return jsonify({'error': 'Invalid or missing fileId (PDF analysis)'}), 400
        
        template = excel_templates[excel_id]
        analysis = document_analysis[file_id]
        
        if not template.get('path') or not Path(template['path']).exists():
            return jsonify({'error': 'Excel template file is missing on disk. Please re-upload the workbook.'}), 400
        
        if not is_zip_file_header(template['path']):
            return jsonify({
                'error': (
                    'Excel template is not a valid .xlsx (ZIP-based) file. '
                    'It may be an older .xls workbook renamed to .xlsx. '
                    'Please "Save As" .xlsx and upload again.'
                )
            }), 400
        
        # Count total questions to process
        total_questions = 0
        workbook_for_count = load_workbook(template['path'])
        marker_cell_regex = re.compile(r'\b([A-Z]\.\d+)\b')
        
        for worksheet in workbook_for_count.worksheets:
            for row in worksheet.iter_rows():
                cell_a = row[0] if len(row) > 0 else None
                if cell_a:
                    raw_a = str(cell_a.value or '').strip()
                    if raw_a and not marker_cell_regex.match(raw_a):
                        total_questions += 1
        
        job_id = create_id('fillJob')
        fill_jobs[job_id] = {
            'status': 'processing',
            'current': 0,
            'total': total_questions,
            'startTime': int(time.time() * 1000),
            'error': None,
            'result': None
        }
        
        # Start background processing (in a real async app, use threading or celery)
        # For simplicity, we'll process synchronously here but you should use threading
        import threading
        thread = threading.Thread(target=process_fill_job, args=(job_id, excel_id, file_id))
        thread.start()
        
        return jsonify({'success': True, 'jobId': job_id, 'total': total_questions})
        
    except Exception as error:
        print(f'Fill Excel error: {error}')
        return jsonify({'error': f'Fill workbook failed: {error}'}), 500


def process_fill_job(job_id, excel_id, file_id):
    """Background job processor for filling Excel workbook."""
    try:
        job = fill_jobs[job_id]
        template = excel_templates[excel_id]
        analysis = document_analysis[file_id]
        
        full_text = analysis.get('content', '')
        
        workbook = load_workbook(template['path'])
        marker_cell_regex = re.compile(r'\b([A-Z]\.\d+)\b')
        
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                cell_a = row[0] if len(row) > 0 else None
                if not cell_a:
                    continue
                
                raw_a = str(cell_a.value or '').strip()
                if not raw_a:
                    continue
                
                # Skip rows that are section markers
                if marker_cell_regex.match(raw_a):
                    continue
                
                cell_b = row[1] if len(row) > 1 else None
                if not cell_b:
                    continue
                
                # Generate answer using Azure OpenAI
                payload = generate_answer_with_azure_openai(
                    marker='',
                    question=raw_a,
                    section_text=full_text
                )
                
                cell_b.value = str(payload.get('answer', 'N/A'))[:250]
                job['current'] += 1
        
        # Save filled workbook
        safe_base = sanitize_filename((analysis.get('filename', 'analysis')).replace('.pdf', '').replace('.PDF', ''))
        out_name = f"{safe_base}_filled.xlsx"
        filled_id = create_id('filled')
        out_path = Path(UPLOAD_FOLDER) / f"{filled_id}-{out_name}"
        workbook.save(out_path)
        
        filled_workbooks[filled_id] = {
            'filename': out_name,
            'path': str(out_path),
            'createdTime': datetime.now().isoformat()
        }
        
        job['status'] = 'complete'
        job['result'] = {
            'filledId': filled_id,
            'filename': out_name,
            'downloadUrl': f'/download-filled-excel/{filled_id}'
        }
        
    except Exception as error:
        print(f'Fill job error: {error}')
        fill_jobs[job_id]['status'] = 'error'
        fill_jobs[job_id]['error'] = str(error)


@app.route('/fill-excel-progress/<job_id>')
def fill_excel_progress(job_id):
    """Get progress of a fill job."""
    job = fill_jobs.get(job_id)
    
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    
    return jsonify({
        'status': job['status'],
        'current': job['current'],
        'total': job['total'],
        'error': job['error'],
        'result': job['result']
    })


@app.route('/export-tables', methods=['POST'])
def export_tables():
    """Export tables to Excel workbook."""
    try:
        data = request.get_json()
        file_id = data.get('fileId')
        
        if not file_id or file_id not in document_analysis:
            return jsonify({'error': 'Invalid or missing fileId'}), 400
        
        analysis = document_analysis[file_id]
        
        if not analysis.get('tables') or len(analysis['tables']) == 0:
            return jsonify({'error': 'No tables found in this document'}), 400
        
        # Create new workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Create worksheet for each table
        for table in analysis['tables']:
            sheet_name = f"Table {table['tableIndex']}"
            worksheet = workbook.create_sheet(sheet_name)
            
            # Add table data
            for row_index, row_data in enumerate(table['data']):
                for col_index, cell_value in enumerate(row_data):
                    cell = worksheet.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
                    
                    # Style header row
                    if row_index == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        
        # Save workbook
        safe_base = sanitize_filename(analysis.get('filename', 'tables').replace('.pdf', '').replace('.PDF', ''))
        out_name = f"{safe_base}_tables.xlsx"
        export_id = create_id('export')
        out_path = Path(UPLOAD_FOLDER) / f"{export_id}-{out_name}"
        workbook.save(out_path)
        
        filled_workbooks[export_id] = {
            'filename': out_name,
            'path': str(out_path),
            'createdTime': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'exportId': export_id,
            'filename': out_name,
            'tableCount': len(analysis['tables'])
        })
        
    except Exception as error:
        print(f'Export tables error: {error}')
        return jsonify({'error': f'Export tables failed: {error}'}), 500


@app.route('/export-single-tab', methods=['POST'])
def export_single_tab():
    """Export tables to single tab."""
    try:
        data = request.get_json()
        file_id = data.get('fileId')
        
        if not file_id or file_id not in document_analysis:
            return jsonify({'error': 'Invalid or missing fileId'}), 400
        
        analysis = document_analysis[file_id]
        
        if not analysis.get('tables') or len(analysis['tables']) == 0:
            return jsonify({'error': 'No tables found in this document'}), 400
        
        # Create new workbook with single worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'All Tables'
        
        current_row = 1
        
        # Add each table vertically
        for table in analysis['tables']:
            # Add table header
            header_cell = worksheet.cell(row=current_row, column=1, value=f"Table {table['tableIndex']}")
            header_cell.font = Font(bold=True, size=12, color='FFFFFFFF')
            header_cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
            
            # Merge cells for table header
            if table['columnCount'] > 1:
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=table['columnCount'])
            
            current_row += 1
            
            # Add table data
            for row_index, row_data in enumerate(table['data']):
                for col_index, cell_value in enumerate(row_data):
                    cell = worksheet.cell(row=current_row, column=col_index + 1, value=cell_value)
                    
                    # Style first row of each table
                    if row_index == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
                
                current_row += 1
            
            # Add spacing between tables
            current_row += 2
        
        # Save workbook
        safe_base = sanitize_filename(analysis.get('filename', 'tables').replace('.pdf', '').replace('.PDF', ''))
        out_name = f"{safe_base}_single_tab.xlsx"
        export_id = create_id('export')
        out_path = Path(UPLOAD_FOLDER) / f"{export_id}-{out_name}"
        workbook.save(out_path)
        
        filled_workbooks[export_id] = {
            'filename': out_name,
            'path': str(out_path),
            'createdTime': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'exportId': export_id,
            'filename': out_name,
            'tableCount': len(analysis['tables'])
        })
        
    except Exception as error:
        print(f'Export single tab error: {error}')
        return jsonify({'error': f'Export single tab failed: {error}'}), 500


@app.route('/download-filled-excel/<filled_id>')
def download_filled_excel(filled_id):
    """Download filled workbook."""
    try:
        entry = filled_workbooks.get(filled_id)
        if not entry or not entry.get('path') or not Path(entry['path']).exists():
            return jsonify({'error': 'Filled workbook not found'}), 404
        
        return send_file(
            entry['path'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=entry['filename']
        )
        
    except Exception as error:
        print(f'Download filled Excel error: {error}')
        return jsonify({'error': 'Failed to download filled workbook'}), 500


@app.route('/transcription/<file_id>')
def get_transcription(file_id):
    """Get document analysis by file ID."""
    analysis = document_analysis.get(file_id)
    
    if not analysis:
        return jsonify({'error': 'Document analysis not found'}), 404
    
    return jsonify(analysis)


@app.route('/transcriptions')
def get_all_transcriptions():
    """Get all document analyses."""
    return jsonify(document_analysis)


@app.route('/chat', methods=['POST'])
def chat():
    """Chat with document content."""
    try:
        data = request.get_json()
        message = data.get('message')
        file_id = data.get('fileId')
        
        if not message:
            return jsonify({'error': 'No message provided'}), 400
        
        # Get document content if fileId is provided
        context = ''
        if file_id and file_id in document_analysis:
            context = document_analysis[file_id]['content']
        
        endpoint_base = normalize_base_url(
            require_any_env(
                ['AZURE_OPENAI_ENDPOINT', 'REACT_APP_AZURE_OPENAI_CHAT_ENDPOINT', 'REACT_APP_AZURE_OPENAI_ENDPOINT'],
                'AZURE_OPENAI_ENDPOINT'
            ),
            'AZURE_OPENAI_ENDPOINT'
        )
        api_key = require_any_env(
            ['AZURE_OPENAI_API_KEY', 'AZURE_OPENAI_KEY', 'REACT_APP_AZURE_OPENAI_CHAT_KEY', 'REACT_APP_AZURE_OPENAI_KEY'],
            'AZURE_OPENAI_API_KEY'
        )
        deployment = require_any_env(
            ['AZURE_OPENAI_DEPLOYMENT', 'REACT_APP_AZURE_OPENAI_DEPLOYMENT', 'REACT_APP_AZURE_OPENAI_DEPLOYMENT_NAME'],
            'AZURE_OPENAI_DEPLOYMENT'
        )
        api_version = get_env('AZURE_OPENAI_API_VERSION') or '2024-02-15-preview'
        
        url = urljoin(
            endpoint_base,
            f"openai/deployments/{deployment}/chat/completions?api-version={api_version}"
        )
        
        # Call Azure OpenAI API
        system_content = (
            f'You are a helpful assistant. Answer questions based on the following document content:\n\n{context}'
            if context else 'You are a helpful assistant.'
        )
        
        response = requests.post(
            url,
            headers={
                'api-key': api_key,
                'Content-Type': 'application/json'
            },
            json={
                'messages': [
                    {'role': 'system', 'content': system_content},
                    {'role': 'user', 'content': message}
                ],
                'max_tokens': 800,
                'temperature': 0.7
            },
            timeout=90
        )
        response.raise_for_status()
        
        ai_response = response.json().get('choices', [{}])[0].get('message', {}).get('content', 'No response generated')
        return jsonify({'response': ai_response})
        
    except Exception as error:
        print(f'Chat error: {error}')
        error_msg = str(error)
        if hasattr(error, 'response') and error.response:
            try:
                error_msg = error.response.json().get('error', {}).get('message', str(error))
            except:
                error_msg = str(error)
        return jsonify({'error': f'Chat failed: {error_msg}'}), 500


@app.route('/files/<file_id>', methods=['DELETE'])
def delete_file(file_id):
    """Delete a file and its data."""
    try:
        # Check if file exists
        if file_id not in document_analysis:
            return jsonify({'error': 'File not found'}), 404
        
        # Delete physical file
        if file_id in file_paths:
            try:
                file_path = Path(file_paths[file_id])
                if file_path.exists():
                    file_path.unlink()
                    print(f'Deleted file: {file_path}')
            except Exception as file_error:
                print(f'Error deleting physical file: {file_error}')
            del file_paths[file_id]
        
        # Remove from document analysis
        del document_analysis[file_id]
        
        return jsonify({
            'success': True,
            'message': 'File deleted successfully',
            'fileId': file_id
        })
        
    except Exception as error:
        print(f'Delete error: {error}')
        return jsonify({'error': f'Failed to delete file: {error}'}), 500


@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error."""
    return jsonify({'error': 'File too large'}), 413


@app.errorhandler(500)
def internal_server_error(error):
    """Handle internal server errors."""
    return jsonify({'error': str(error)}), 500


if __name__ == '__main__':
    # Rebuild file paths mapping from existing files
    rebuild_file_paths_mapping()
    
    # Start Flask server
    port = int(os.getenv('PORT', 8080))
    print(f'Server running on http://localhost:{port}')
    print('Using in-memory storage for document analysis')
    
    app.run(host='0.0.0.0', port=port, debug=True)
